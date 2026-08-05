#!/usr/bin/env python3
"""
Construit les données géographiques de la page NET à partir du fichier Excel maître.

Entrée :
- feuille "artistes"
- colonnes attendues : ville_region, country_code
- colonne artiste détectée parmi : nom, artiste, artist

Sorties :
- app/data/net/geography.json
- app/data/net/geography-review.json
- app/data/net/geography-unmatched.json

Le script conserve les correspondances déjà validées et n'interroge GeoNames
que pour les lieux nouveaux ou non résolus.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "\nERREUR : openpyxl n'est pas installé.\n"
        "Exécute d'abord :\n"
        "  py -m pip install openpyxl\n",
        file=sys.stderr,
    )
    raise SystemExit(1)


GEONAMES_URL = "https://secure.geonames.org/searchJSON"
DEFAULT_OUTPUT_DIR = Path("app/data/net")
API_DELAY_SECONDS = 1.1

# Projection equirectangulaire standard.
# Ces valeurs pourront être ajustées plus tard si le dessin de la carte
# possède des marges internes particulières.
MAP_X_MIN = 0.0
MAP_X_MAX = 100.0
MAP_Y_MIN = 0.0
MAP_Y_MAX = 100.0


@dataclass(frozen=True)
class SourcePlace:
    source_label: str
    country_code: str
    artists: tuple[str, ...]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize_key(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = "".join(char for char in value if not unicodedata.combining(char))
    value = value.casefold()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def slugify_place(source_label: str, country_code: str) -> str:
    base = normalize_key(source_label).replace(" ", "-")
    country = country_code.lower()
    return f"{country}-{base}".strip("-")


def read_env_value(project_root: Path, key: str) -> str | None:
    env_path = project_root / ".env.local"
    if not env_path.exists():
        return os.getenv(key)

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        env_key, env_value = line.split("=", 1)
        if env_key.strip() != key:
            continue

        return env_value.strip().strip('"').strip("'")

    return os.getenv(key)


def find_project_root(start: Path) -> Path:
    candidates = [start.resolve(), *start.resolve().parents]
    for candidate in candidates:
        if (candidate / "app").exists() and (
            (candidate / "package.json").exists() or (candidate / ".env.local").exists()
        ):
            return candidate
    return start.resolve()


def find_excel(project_root: Path, explicit_path: str | None) -> Path:
    if explicit_path:
        path = Path(explicit_path)
        if not path.is_absolute():
            path = project_root / path
        if not path.exists():
            raise FileNotFoundError(f"Fichier Excel introuvable : {path}")
        return path

    patterns = (
        "Metaliciel_Base_Maitre*.xlsx",
        "Métaliciel_Base_Maitre*.xlsx",
        "*Base*Maitre*.xlsx",
        "*Base*Maître*.xlsx",
    )

    matches: list[Path] = []
    ignored_parts = {"node_modules", ".next", ".git"}

    for pattern in patterns:
        for path in project_root.rglob(pattern):
            if any(part in ignored_parts for part in path.parts):
                continue
            if path.name.startswith("~$"):
                continue
            matches.append(path)

    unique_matches = sorted(set(matches), key=lambda item: item.stat().st_mtime, reverse=True)

    if not unique_matches:
        raise FileNotFoundError(
            "Aucun fichier Excel maître trouvé. "
            "Relance avec --excel \"chemin/vers/le-fichier.xlsx\"."
        )

    if len(unique_matches) > 1:
        print("Plusieurs fichiers Excel trouvés. Le plus récent sera utilisé :")
        for item in unique_matches[:5]:
            print(f"  - {item}")

    return unique_matches[0]


def find_sheet_name(workbook: Any, expected: str) -> str:
    expected_key = normalize_key(expected)
    for sheet_name in workbook.sheetnames:
        if normalize_key(sheet_name) == expected_key:
            return sheet_name
    raise KeyError(
        f"Feuille '{expected}' introuvable. Feuilles disponibles : "
        + ", ".join(workbook.sheetnames)
    )


def detect_headers(sheet: Any) -> tuple[int, dict[str, int]]:
    required_groups = {
        "place": {"ville region", "ville_region", "ville", "region ville pays"},
        "country": {"country code", "country_code", "code pays", "iso", "iso2"},
    }

    for row_index in range(1, min(sheet.max_row, 25) + 1):
        headers: dict[str, int] = {}

        for column_index in range(1, sheet.max_column + 1):
            raw_value = sheet.cell(row=row_index, column=column_index).value
            key = normalize_key(normalize_text(raw_value))
            if key:
                headers[key] = column_index

        has_place = any(key in headers for key in required_groups["place"])
        has_country = any(key in headers for key in required_groups["country"])

        if has_place and has_country:
            return row_index, headers

    raise KeyError(
        "Impossible de trouver une ligne d'en-têtes contenant "
        "'ville_region' et 'country_code'."
    )


def find_column(headers: dict[str, int], aliases: Iterable[str]) -> int | None:
    normalized_aliases = [normalize_key(alias) for alias in aliases]

    for alias in normalized_aliases:
        if alias in headers:
            return headers[alias]

    return None


def read_source_places(excel_path: Path) -> list[SourcePlace]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet_name = find_sheet_name(workbook, "artistes")
    sheet = workbook[sheet_name]

    header_row, headers = detect_headers(sheet)

    place_column = find_column(
        headers,
        ("ville_region", "ville région", "ville", "region_ville_pays"),
    )
    country_column = find_column(
        headers,
        ("country_code", "country code", "code pays", "iso", "iso2"),
    )
    artist_column = find_column(
        headers,
        ("nom", "artiste", "artist", "nom artiste", "nom_artiste"),
    )

    if place_column is None or country_column is None:
        raise KeyError(
            "Les colonnes 'ville_region' et 'country_code' sont obligatoires."
        )

    grouped_artists: dict[tuple[str, str], set[str]] = defaultdict(set)
    display_labels: dict[tuple[str, str], str] = {}

    for row_index in range(header_row + 1, sheet.max_row + 1):
        source_label = normalize_text(sheet.cell(row=row_index, column=place_column).value)
        country_code = normalize_text(
            sheet.cell(row=row_index, column=country_column).value
        ).upper()

        if not source_label or not country_code:
            continue

        artist_name = ""
        if artist_column is not None:
            artist_name = normalize_text(
                sheet.cell(row=row_index, column=artist_column).value
            )

        key = (normalize_key(source_label), country_code)
        display_labels.setdefault(key, source_label)

        if artist_name:
            grouped_artists[key].add(artist_name)

    places = [
        SourcePlace(
            source_label=display_labels[key],
            country_code=key[1],
            artists=tuple(sorted(artists, key=str.casefold)),
        )
        for key, artists in grouped_artists.items()
    ]

    return sorted(
        places,
        key=lambda place: (place.country_code, normalize_key(place.source_label)),
    )


def load_json_list(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []

    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as error:
        raise ValueError(f"JSON invalide dans {path}: {error}") from error

    if not isinstance(value, list):
        raise ValueError(f"Le fichier {path} doit contenir une liste JSON.")

    return [item for item in value if isinstance(item, dict)]


def existing_entries_by_key(paths: Iterable[Path]) -> dict[tuple[str, str], dict[str, Any]]:
    entries: dict[tuple[str, str], dict[str, Any]] = {}

    for path in paths:
        for item in load_json_list(path):
            source_label = normalize_text(item.get("sourceLabel"))
            country_code = normalize_text(item.get("countryCode")).upper()

            if source_label and country_code:
                entries[(normalize_key(source_label), country_code)] = item

    return entries


def geonames_request(
    username: str,
    query: str,
    country_code: str,
    feature_class: str,
) -> list[dict[str, Any]]:
    parameters = {
        "q": query,
        "country": country_code,
        "featureClass": feature_class,
        "maxRows": 10,
        "lang": "fr",
        "style": "FULL",
        "type": "json",
        "username": username,
    }

    url = f"{GEONAMES_URL}?{urllib.parse.urlencode(parameters)}"
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "SITE-IM-NET-Geocoder/1.0"},
    )

    try:
        with urllib.request.urlopen(request, timeout=25) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"GeoNames HTTP {error.code}: {detail}") from error
    except urllib.error.URLError as error:
        raise RuntimeError(f"Connexion GeoNames impossible : {error}") from error

    if "status" in payload:
        status = payload["status"]
        raise RuntimeError(
            f"GeoNames a renvoyé une erreur : "
            f"{status.get('message', status)}"
        )

    values = payload.get("geonames", [])
    if not isinstance(values, list):
        return []

    return [value for value in values if isinstance(value, dict)]


def candidate_names(candidate: dict[str, Any]) -> set[str]:
    values = {
        normalize_key(normalize_text(candidate.get("name"))),
        normalize_key(normalize_text(candidate.get("toponymName"))),
        normalize_key(normalize_text(candidate.get("asciiName"))),
    }

    alternate_names = candidate.get("alternateNames", [])
    if isinstance(alternate_names, list):
        for alternate in alternate_names:
            if isinstance(alternate, dict):
                values.add(normalize_key(normalize_text(alternate.get("name"))))

    return {value for value in values if value}


def score_candidate(
    candidate: dict[str, Any],
    source_label: str,
    primary_name: str,
    administrative_hint: str,
) -> float:
    source_key = normalize_key(source_label)
    primary_key = normalize_key(primary_name)
    hint_key = normalize_key(administrative_hint)
    names = candidate_names(candidate)

    score = 0.0

    if primary_key in names:
        score += 60
    elif any(primary_key and primary_key in name for name in names):
        score += 38

    if source_key in names:
        score += 15

    admin_values = [
        normalize_key(normalize_text(candidate.get("adminName1"))),
        normalize_key(normalize_text(candidate.get("adminName2"))),
        normalize_key(normalize_text(candidate.get("adminName3"))),
    ]

    if hint_key:
        if hint_key in admin_values:
            score += 24
        elif any(hint_key in value or value in hint_key for value in admin_values if value):
            score += 13

    feature_class = normalize_text(candidate.get("fcl"))
    feature_code = normalize_text(candidate.get("fcode"))

    if feature_class == "P":
        score += 12
    elif feature_class == "A":
        score += 3

    if feature_code in {"PPLC", "PPLA", "PPLA2", "PPLA3", "PPLA4", "PPL"}:
        score += 5

    try:
        population = int(candidate.get("population") or 0)
    except (TypeError, ValueError):
        population = 0

    if population > 0:
        score += min(10.0, math.log10(population + 1) * 1.5)

    return round(score, 2)


def split_source_label(source_label: str) -> tuple[str, str]:
    parts = [part.strip() for part in source_label.split(",") if part.strip()]

    if not parts:
        return source_label, ""

    primary_name = parts[0]
    administrative_hint = ", ".join(parts[1:])
    return primary_name, administrative_hint


def search_candidates(
    username: str,
    source_label: str,
    country_code: str,
) -> list[dict[str, Any]]:
    primary_name, administrative_hint = split_source_label(source_label)

    searches: list[tuple[str, str]] = [
        (source_label, "P"),
        (primary_name, "P"),
    ]

    # Certains lieux du fichier maître représentent une région plutôt qu'une ville.
    searches.extend(
        [
            (source_label, "A"),
            (primary_name, "A"),
        ]
    )

    collected: dict[str, dict[str, Any]] = {}

    for query, feature_class in searches:
        if not query:
            continue

        candidates = geonames_request(
            username=username,
            query=query,
            country_code=country_code,
            feature_class=feature_class,
        )

        for candidate in candidates:
            geoname_id = normalize_text(candidate.get("geonameId"))
            if geoname_id:
                collected[geoname_id] = candidate

        time.sleep(API_DELAY_SECONDS)

    ranked: list[dict[str, Any]] = []

    for candidate in collected.values():
        enriched = dict(candidate)
        enriched["_score"] = score_candidate(
            candidate,
            source_label=source_label,
            primary_name=primary_name,
            administrative_hint=administrative_hint,
        )
        ranked.append(enriched)

    ranked.sort(
        key=lambda item: (
            float(item.get("_score", 0)),
            int(item.get("population") or 0),
        ),
        reverse=True,
    )

    return ranked


def map_coordinates(latitude: float, longitude: float) -> tuple[float, float]:
    raw_x = (longitude + 180.0) / 360.0
    raw_y = (90.0 - latitude) / 180.0

    map_x = MAP_X_MIN + raw_x * (MAP_X_MAX - MAP_X_MIN)
    map_y = MAP_Y_MIN + raw_y * (MAP_Y_MAX - MAP_Y_MIN)

    return round(map_x, 4), round(map_y, 4)


def serialize_candidate(candidate: dict[str, Any]) -> dict[str, Any]:
    return {
        "geonameId": candidate.get("geonameId"),
        "name": candidate.get("name"),
        "toponymName": candidate.get("toponymName"),
        "countryCode": candidate.get("countryCode"),
        "adminName1": candidate.get("adminName1"),
        "adminName2": candidate.get("adminName2"),
        "featureClass": candidate.get("fcl"),
        "featureCode": candidate.get("fcode"),
        "latitude": candidate.get("lat"),
        "longitude": candidate.get("lng"),
        "population": candidate.get("population"),
        "score": candidate.get("_score"),
    }


def matched_entry(place: SourcePlace, candidate: dict[str, Any]) -> dict[str, Any]:
    latitude = float(candidate["lat"])
    longitude = float(candidate["lng"])
    map_x, map_y = map_coordinates(latitude, longitude)

    return {
        "id": slugify_place(place.source_label, place.country_code),
        "sourceLabel": place.source_label,
        "countryCode": place.country_code,
        "matchedName": candidate.get("name"),
        "toponymName": candidate.get("toponymName"),
        "adminName1": candidate.get("adminName1"),
        "adminName2": candidate.get("adminName2"),
        "geonameId": candidate.get("geonameId"),
        "featureClass": candidate.get("fcl"),
        "featureCode": candidate.get("fcode"),
        "latitude": latitude,
        "longitude": longitude,
        "mapX": map_x,
        "mapY": map_y,
        "artistCount": len(place.artists),
        "artists": list(place.artists),
        "status": "matched",
        "confidenceScore": candidate.get("_score"),
    }


def review_entry(
    place: SourcePlace,
    candidates: list[dict[str, Any]],
    reason: str,
) -> dict[str, Any]:
    return {
        "id": slugify_place(place.source_label, place.country_code),
        "sourceLabel": place.source_label,
        "countryCode": place.country_code,
        "artistCount": len(place.artists),
        "artists": list(place.artists),
        "status": "review",
        "reason": reason,
        "candidates": [serialize_candidate(candidate) for candidate in candidates[:5]],
    }


def unmatched_entry(place: SourcePlace, reason: str) -> dict[str, Any]:
    return {
        "id": slugify_place(place.source_label, place.country_code),
        "sourceLabel": place.source_label,
        "countryCode": place.country_code,
        "artistCount": len(place.artists),
        "artists": list(place.artists),
        "status": "unmatched",
        "reason": reason,
    }


def refresh_existing_entry(
    existing: dict[str, Any],
    place: SourcePlace,
) -> dict[str, Any]:
    refreshed = dict(existing)
    refreshed["artistCount"] = len(place.artists)
    refreshed["artists"] = list(place.artists)
    return refreshed


def write_json(path: Path, data: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Génère les données géographiques de la page NET."
    )
    parser.add_argument(
        "--excel",
        help="Chemin du fichier Excel maître. Recherche automatique si omis.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="Dossier de sortie relatif à la racine du projet.",
    )
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="Relance GeoNames même pour les lieux déjà connus.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Limite le nombre de nouveaux lieux traités pour un test.",
    )
    args = parser.parse_args()

    project_root = find_project_root(Path.cwd())
    username = read_env_value(project_root, "GEONAMES_USERNAME")

    if not username:
        print(
            "ERREUR : GEONAMES_USERNAME est absent de .env.local.",
            file=sys.stderr,
        )
        return 1

    excel_path = find_excel(project_root, args.excel)
    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = project_root / output_dir

    geography_path = output_dir / "geography.json"
    review_path = output_dir / "geography-review.json"
    unmatched_path = output_dir / "geography-unmatched.json"

    print(f"\nProjet        : {project_root}")
    print(f"Excel utilisé : {excel_path}")
    print(f"GeoNames      : {username}")
    print(f"Sortie        : {output_dir}\n")

    places = read_source_places(excel_path)
    existing = existing_entries_by_key(
        (geography_path, review_path, unmatched_path)
    )

    matched: list[dict[str, Any]] = []
    review: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []

    processed_new = 0

    for index, place in enumerate(places, start=1):
        key = (normalize_key(place.source_label), place.country_code)
        old_entry = existing.get(key)

        if old_entry and not args.refresh:
            refreshed = refresh_existing_entry(old_entry, place)
            status = refreshed.get("status")

            if status == "matched":
                matched.append(refreshed)
            elif status == "review":
                review.append(refreshed)
            else:
                unmatched.append(refreshed)

            print(
                f"[{index:>3}/{len(places)}] conservé : "
                f"{place.source_label} ({place.country_code})"
            )
            continue

        if args.limit and processed_new >= args.limit:
            unmatched.append(
                unmatched_entry(place, "Non traité : limite de test atteinte.")
            )
            continue

        processed_new += 1
        print(
            f"[{index:>3}/{len(places)}] recherche : "
            f"{place.source_label} ({place.country_code})"
        )

        try:
            candidates = search_candidates(
                username=username,
                source_label=place.source_label,
                country_code=place.country_code,
            )
        except RuntimeError as error:
            unmatched.append(unmatched_entry(place, str(error)))
            print(f"       ERREUR : {error}")
            continue

        if not candidates:
            unmatched.append(
                unmatched_entry(place, "Aucun résultat GeoNames.")
            )
            print("       aucun résultat")
            continue

        best = candidates[0]
        best_score = float(best.get("_score", 0))
        second_score = (
            float(candidates[1].get("_score", 0))
            if len(candidates) > 1
            else 0.0
        )
        score_gap = best_score - second_score

        # Validation automatique :
        # - score élevé ;
        # - avantage net sur le deuxième candidat.
        if best_score >= 70 and (len(candidates) == 1 or score_gap >= 8):
            matched.append(matched_entry(place, best))
            print(
                f"       trouvé : {best.get('name')} / "
                f"{best.get('adminName1', '')} — score {best_score}"
            )
        else:
            reason = (
                f"Correspondance à vérifier : meilleur score {best_score}, "
                f"écart {round(score_gap, 2)}."
            )
            review.append(review_entry(place, candidates, reason))
            print(f"       à vérifier — {reason}")

    matched.sort(
        key=lambda item: (
            item.get("countryCode", ""),
            normalize_key(str(item.get("sourceLabel", ""))),
        )
    )
    review.sort(
        key=lambda item: (
            item.get("countryCode", ""),
            normalize_key(str(item.get("sourceLabel", ""))),
        )
    )
    unmatched.sort(
        key=lambda item: (
            item.get("countryCode", ""),
            normalize_key(str(item.get("sourceLabel", ""))),
        )
    )

    write_json(geography_path, matched)
    write_json(review_path, review)
    write_json(unmatched_path, unmatched)

    print("\nTerminé.")
    print(f"  Correspondances validées : {len(matched)}")
    print(f"  À vérifier               : {len(review)}")
    print(f"  Non trouvées             : {len(unmatched)}")
    print(f"\nFichiers générés dans : {output_dir}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
